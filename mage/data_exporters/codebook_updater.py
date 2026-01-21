import os
import io
import openpyxl
from datetime import datetime

from utils.helper_functions import get_previous_week_window, s3

if 'data_exporter' not in globals():
    from mage_ai.data_preparation.decorators import data_exporter


@data_exporter
def export_data(data, *args, **kwargs):
    execution_date = kwargs.get('execution_date')

    if execution_date is None:
        execution_date = datetime.now()

    start, end = get_previous_week_window(execution_date)

    bucket_name = os.getenv('WASABI_BUCKET')
    prefix = os.getenv('WASABI_PREFIX')

    file_key = f"{prefix}/agmarknet_codebook.xlsx"

    s3_client = s3()

    try:
        print(f"Downloading {file_key} from Wasabi...")
        
        response = s3_client.get_object(Bucket=bucket_name, Key=file_key)
        
        file_content = response['Body'].read()

        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        
        if 'metadata information' not in wb.sheetnames:
            print(f"Error: Sheet 'metadata information' not found in {file_key}")
            
        ws = wb['metadata information']

        retrieval_val = execution_date.strftime('%Y-%m-%d')
        updated_val = end.strftime('%Y-%m-%d')

        found_retrieval = False
        found_updated = False

        for row in range(1, ws.max_row + 1):
            cell_a = ws.cell(row=row, column=1).value
            
            if cell_a == "Data Retrieval Date":
                ws.cell(row=row, column=2).value = retrieval_val
                found_retrieval = True
            
            elif cell_a == "Data Last Updated":
                ws.cell(row=row, column=2).value = updated_val
                found_updated = True

        if not (found_retrieval and found_updated):
            print("Warning: Could not find one or both labels in Column A.")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        s3_client.put_object(
            Bucket=bucket_name,
            Key=file_key,
            Body=output.getvalue()
        )

        print(f"Successfully updated metadata in {file_key} and uploaded to Wasabi.")
        return data
        
    except Exception as e:
        print(f"Failed to update codebook: {str(e)}")